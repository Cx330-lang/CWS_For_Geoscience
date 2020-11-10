#coding=gbk
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='gb18030')

import openpyxl
import pandas as pd
import jieba
import numpy as np
from pyltp import Segmentor
from stanfordcorenlp import StanfordCoreNLP
import pynlpir
import thulac
import pkuseg
from pyhanlp import HanLP
# import fool

# wb.save(excel_path)

class DataProcess:

    def __init__(self, txt_path, excel_path):
        self.txt_path = txt_path
        self.excel_path = excel_path

        self.wb = openpyxl.load_workbook(self.excel_path)
        self.ws = self.wb[self.wb.sheetnames[0]]

    def load_dataset(self, max_sent_len=256):
        # f = open(self.txt_path, 'r', encoding='utf-8')
        # train_lines = f.read().strip().split('\n')
        # f.close()
        df = pd.read_excel(self.excel_path)

        train_lines = df['word_sents']

        train_lines = [line.strip().split() for line in train_lines]

        # print(train_lines)
        # print(len(train_lines))

        train_pruned_lines = []

        # prun data set
        for train_line in train_lines:
            current_length = 0
            line = []
            for word in train_line:
                line.append(word)
                current_length += len(word)
                if current_length > max_sent_len:
                    train_pruned_lines.append(line.copy())
                    current_length = 0
                    line = []
            if len(line) == 0:
                continue
            train_pruned_lines.append(line.copy())

        # print(train_pruned_lines)

        return train_pruned_lines

    def save_data(self):
    #     train_sents, train_tags = [], []

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb[wb.sheetnames[0]]
        ws['B1'] = 'sentences'
        ws['C1'] = 'tags'

        train_pruned_lines = self.load_dataset()
        for i, line in enumerate(train_pruned_lines, start=2):
            cur_sent = ''.join(line)
            pos = 0
            cur_tag = [0] * (len(cur_sent))
            if len(cur_sent) == 0:
                # print(i,line)
                continue
            for word in line:
                if len(word) == 1:
                    # single word
                    cur_tag[pos] = 'S'
                    pos += 1
                else:
                    # more than one word
                    cur_tag[pos] = 'B'
                    cur_tag[pos + len(word) - 1] = "E"
                    if len(word) > 2:
                        cur_tag[pos + 1:pos + len(word) - 1] = ['M'] * (len(word) - 2)
                    pos = pos + len(word)


            ws['B'+str(i)] = cur_sent
            ws['C' + str(i)] = ''.join(cur_tag)
            # train_sents.append(cur_sent)
            # train_tags.append(','.join(cur_tag))

        wb.save(self.excel_path)

    def save_to_excel(self, lines):
        self.ws['N1'] = 'hanlp_cut_sents'
        self.ws['O1'] = 'hanlp_cut_tags'

        for i, line in enumerate(lines, start=2):
            self.ws['N'+str(i)] = line
            line = line.strip().split()
            cur_sent = ''.join(line)
            pos = 0
            cur_tag = [0] * (len(cur_sent))
            if len(cur_sent) == 0:
                # print(i,line)
                continue
            for word in line:
                if len(word) == 1:
                    # single word
                    cur_tag[pos] = 'S'
                    pos += 1
                else:
                    # more than one word
                    cur_tag[pos] = 'B'
                    cur_tag[pos + len(word) - 1] = "E"
                    if len(word) > 2:
                        cur_tag[pos + 1:pos + len(word) - 1] = ['M'] * (len(word) - 2)
                    pos = pos + len(word)

            self.ws['O' + str(i)] = ''.join(cur_tag)

        self.wb.save(self.excel_path)


    def jieba_cut(self):
        df = pd.read_excel(self.excel_path)
        sentences = df['sentences']
        lines = []
        for sentence in sentences:
            # print(sentence)
            # print(type(sentence))
            if sentence is not np.nan:
                cut = jieba.cut(sentence)
                lines.append(' '.join(cut))
        # print(lines)

        self.save_to_excel(lines)

    def ltp_cut(self):
        model_path = r'E:\BaiduNetdiskDownload\3.4.0\ltp_data_v3.4.0\cws.model'
        segmentor = Segmentor()
        segmentor.load(model_path)  # 加载模型
        # words = segmentor.segment(sentence)
        df = pd.read_excel(self.excel_path)
        sentences = df['sentences']
        lines = []
        for sentence in sentences:
            if sentence is not np.nan:
                cut = segmentor.segment(sentence)
                lines.append(' '.join(cut))
        # print(lines)
        self.save_to_excel(lines)

    def stanford_cut(self):
        path = r'D:\Tools\Package\stanford-corenlp-full-2018-02-27'

        lines = []
        nlp = StanfordCoreNLP(path, lang='zh')
        df = pd.read_excel(self.excel_path)
        sentences = df['sentences']
        for sentence in sentences:
            if sentence is not np.nan:
                # print(sentence)
                    cut = nlp.word_tokenize(sentence)
                    lines.append(' '.join(cut))
        # print(lines)
        self.save_to_excel(lines)

    def ictclas_cut(self):
        pynlpir.open()
        lines = []
        df = pd.read_excel(self.excel_path)
        sentences = df['sentences']
        for sentence in sentences:
            if sentence is not np.nan:
                # print(sentence)
                str_sents = ''
                cuts = pynlpir.segment(sentence)
                for cut in cuts:
                    str_sents += cut[0] + ' '
                lines.append(str_sents)
        # print(lines)
        pynlpir.close()

        self.save_to_excel(lines)

    def thu_cut(self):
        thul_cut = thulac.thulac(seg_only=True)
        lines = []
        df = pd.read_excel(self.excel_path)
        sentences = df['sentences']
        for sentence in sentences:
            if sentence is not np.nan:
                str_sents = ''
                # print(sentence)
                cuts = thul_cut.cut(sentence)
                for cut in cuts:
                    str_sents += cut[0] + ' '
                lines.append(str_sents)
                # print(cut)
        # print(lines)
        self.save_to_excel(lines)

    def hanlp_cut(self):
        lines = []
        df = pd.read_excel(self.excel_path)
        sentences = df['sentences']
        for sentence in sentences:
            if sentence is not np.nan:
                # print(sentence)
                cuts = HanLP.segment(sentence)
                lines.append(' '.join(cut.word for cut in cuts))

        # print(lines)
        self.save_to_excel(lines)

    def pkuseg_cut(self):
        model_path = r'D:\Anaconda\python\envs\spider\Lib\site-packages\pkuseg\models\default\tourism'
        seg = pkuseg.pkuseg(model_name=model_path)
        lines = []
        df = pd.read_excel(self.excel_path)
        sentences = df['sentences']
        for sentence in sentences:
            if sentence is not np.nan:
                # print(sentence)
                cut = seg.cut(sentence)
                lines.append(' '.join(cut))

        # print(lines)
        self.save_to_excel(lines)

def main():
    txt_path = r'D:\Pycharm\Project\data_analyze\Data_processing\data_2\地质数据.txt'
    excel_path = r'D:\Pycharm\Project\data_analyze\Data_processing\data_2\data.xlsx'

    dp = DataProcess(txt_path, excel_path)

    # 载入数据集
    # dp.load_dataset()

    # 数据存取
    # dp.save_data()

    # jieba分词
    # dp.jieba_cut()

    # ltp分词
    # dp.ltp_cut()

    # nltk分词
    # dp.stanford_cut()

    #ictclas 分词
    # dp.ictclas_cut()

    #thulac 分词
    # dp.thu_cut()

    #pkuseg 分词
    # dp.pkuseg_cut()

    #hanlp分词
    dp.hanlp_cut()


if __name__ == '__main__':
    main()